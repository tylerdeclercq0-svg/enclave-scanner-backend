"""
Diligence Tracker export -- server-side XLSX generation.

Design decision (2026-07-06): the frontend computes each parcel's
verification checklist via `buildVerificationChecklist(r)` in
`web/index.html` and POSTs the pre-computed checklist alongside the raw
row data. This module just formats the payload into a styled openpyxl
workbook -- it does NOT re-derive checklist status from the row, so what
gets exported always matches what the UI showed. The alternative
(recomputing statuses in Python) would drift the moment the JS checklist
logic changed, silently.

Later phase (per Tyler, 2026-07-06): a per-parcel PDF checklist. This
module is deliberately structured so the same checklist payload can be
handed to a separate PDF renderer -- see `_normalize_checklist_columns`
which handles the "different parcels expose different checklist items"
problem in a way that a PDF renderer will also need.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Excel's standard "conditional-formatting" color set (Home > Cell Styles
# > Good/Bad/Neutral use these exact fills), so opening the export in
# Excel matches the visual language a real estate analyst already knows.
_STATUS_FILLS = {
    "pass":   PatternFill("solid", fgColor="C6EFCE"),  # light green
    "fail":   PatternFill("solid", fgColor="FFC7CE"),  # light red
    "est":    PatternFill("solid", fgColor="FFEB9C"),  # light yellow
    "manual": PatternFill("solid", fgColor="D9D9D9"),  # light gray
}
_STATUS_LABELS = {
    "pass": "Automated pass",
    "fail": "Automated fail",
    "est": "Estimated (verify)",
    "manual": "Manual required",
}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1E2A24")  # --ink from the app palette
_STATUS_FONT = Font(bold=True)

# One canonical export as of 2026-07-06 -- previously the CSV export and
# diligence tracker were two separate downloads; merged into a single
# workbook so an analyst has ONE living tracker file per selection with
# every raw field, every checklist status, watch-list tier, and blank
# confirm/date columns per checklist item.
#
# Column groups, ordered left-to-right so the most decision-relevant
# info is visible without scrolling. Freeze panes below covers this
# left set:
#   1. IDENTIFIERS (frozen columns) -- parcel_id, county, acres, owner
#   2. PATHWAY & SCORING -- qual %, pathways, confidence tier, score
#   3. PROPERTY/SITE data -- use code, owner_2, jurisdiction, water/sewer,
#      FLUM character, surrounding density, ownership signals, centroid
#   4. ZIP SECTION -- zcta5 (from the coverage_ledger flow)
#   5. VERIFICATION CHECKLIST -- see checklist_labels; three columns per
#      item (Status color-coded / Confirmed by / Date confirmed)
#   6. NOTES -- free-text (blank)
_CORE_COLUMNS = [
    # ---- 1. IDENTIFIERS (frozen) ----
    ("parcel_id", "Parcel ID", 22),
    ("county_id", "County", 12),
    ("acreage", "Acres", 8),
    ("owner_name", "Owner", 30),
    # ---- 2. PATHWAY & SCORING (decision-relevant, visible on load) ----
    ("pct_perimeter_qualifying", "Qual. perimeter %", 12),
    ("pathways_str", "Pathways matched", 18),
    ("confidence_tier", "Confidence tier", 14),
    ("attractiveness_score", "Score /100", 10),
    # ---- 3. PROPERTY / SITE data ----
    ("owner_name_2", "Co-owner", 24),
    ("use_code", "Use code", 10),
    ("jurisdiction", "Jurisdiction", 16),
    ("sold_since_2025", "Sold since 1/1/2025", 14),
    ("single_owner_signal", "Single owner signal", 14),
    ("water_source", "Water source (est.)", 22),
    ("wastewater_method", "Wastewater (est.)", 22),
    ("water_sewer_confidence", "Water/sewer conf.", 14),
    ("flum_character", "Own FLUM character", 20),
    ("surrounding_density", "Surrounding density", 16),
    ("site_address", "Site address", 30),
    ("centroid_lat", "Centroid lat", 12),
    ("centroid_lon", "Centroid lon", 12),
    ("acreage_source", "Acreage source", 18),
    # ---- 4. ZIP SECTION ----
    ("zcta5", "ZCTA / ZIP section", 12),
]

# Number of leftmost columns to freeze (in addition to row 1). Set so
# the identifier group stays visible while scrolling through the
# checklist columns to the right.
_FROZEN_COLUMN_COUNT = 4


def _normalize_checklist_columns(payload_rows: list[dict[str, Any]]) -> list[str]:
    """
    Build the union of checklist item labels across all selected parcels,
    preserving first-seen order. Different parcels can expose different
    items (e.g. Osceola gets an extra Option 5 row that other counties
    don't) -- we don't want to silently drop the Osceola-specific columns
    just because Pasco doesn't have them.
    """
    seen: list[str] = []
    seen_set: set[str] = set()
    for row in payload_rows:
        for item in row.get("checklist") or []:
            label = item.get("label")
            if label and label not in seen_set:
                seen.append(label)
                seen_set.add(label)
    return seen


def build_diligence_tracker_xlsx(payload_rows: list[dict[str, Any]]) -> bytes:
    """
    Build a formatted diligence-tracker workbook from client-supplied rows
    + checklist payload. Returns raw .xlsx bytes.

    Layout:
      Row 1: bold headers on ink background.
      Column A: parcel_id (frozen so identifiers stay visible when
                scrolling through the checklist columns).
      Row 1 also frozen.
      Columns B-G: county / acres / owner / site address / qualifying %
                   / pathways string.
      Then, for each verification-checklist item (union across selection),
      three columns: "<label> - Status" (color-coded), "<label> - Confirmed
      by" (blank), "<label> - Date confirmed" (blank).
      Final column: "Notes" (blank).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Diligence Tracker"

    checklist_labels = _normalize_checklist_columns(payload_rows)

    # ---- headers ----
    headers: list[tuple[str, int]] = [(label, width) for _, label, width in _CORE_COLUMNS]
    for label in checklist_labels:
        headers.append((f"{label} - Status", 22))
        headers.append((f"{label} - Confirmed by", 18))
        headers.append((f"{label} - Date confirmed", 14))
    headers.append(("Notes", 40))

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32

    # ---- data rows ----
    # Fills for the confidence_tier cell so the primary result column
    # is immediately readable at a glance. brass-dim for confident/possible,
    # yellow for the new watch tier, gray for unlikely.
    tier_fills = {
        "confident": PatternFill("solid", fgColor="C6EFCE"),
        "possible":  PatternFill("solid", fgColor="EEE3C2"),
        "watch":     PatternFill("solid", fgColor="FFEB9C"),
        "unlikely":  PatternFill("solid", fgColor="D9D9D9"),
    }

    for row_idx, row in enumerate(payload_rows, start=2):
        core_values = {
            # 1. Identifiers
            "parcel_id": row.get("parcel_id"),
            "county_id": row.get("county_id"),
            "acreage": row.get("acreage"),
            "owner_name": row.get("owner_name"),
            # 2. Pathway & scoring
            "pct_perimeter_qualifying": row.get("pct_perimeter_qualifying"),
            "pathways_str": ", ".join(f"Option {p}" for p in (row.get("likely_pathways") or [])) or "(none)",
            "confidence_tier": (row.get("confidence_tier") or "").capitalize() or "-",
            "attractiveness_score": row.get("attractiveness_score"),
            # 3. Property / site
            "owner_name_2": row.get("owner_name_2"),
            "use_code": row.get("use_code"),
            "jurisdiction": row.get("jurisdiction"),
            "sold_since_2025": (
                "Yes" if row.get("sold_since_2025") is True
                else "No" if row.get("sold_since_2025") is False
                else "Unknown"
            ),
            "single_owner_signal": (
                "No co-owner on record" if row.get("single_owner_signal") is True
                else "Co-owner recorded" if row.get("single_owner_signal") is False
                else "Unknown"
            ),
            "water_source": row.get("water_source"),
            "wastewater_method": row.get("wastewater_method"),
            "water_sewer_confidence": row.get("water_sewer_confidence"),
            "flum_character": row.get("flum_character"),
            "surrounding_density": row.get("surrounding_density"),
            "site_address": row.get("site_address"),
            "centroid_lat": row.get("centroid_lat"),
            "centroid_lon": row.get("centroid_lon"),
            "acreage_source": row.get("acreage_source"),
            # 4. ZIP section
            "zcta5": row.get("zcta5"),
        }
        for col_idx, (key, _, _) in enumerate(_CORE_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=core_values.get(key))
            # Color-fill the confidence_tier cell to mirror the in-app tier
            # badges, so the tracker matches the UI at a glance.
            if key == "confidence_tier":
                tier_key = (row.get("confidence_tier") or "").lower()
                if tier_key in tier_fills:
                    cell.fill = tier_fills[tier_key]
                    cell.font = _STATUS_FONT

        # Index this row's checklist by label so we can look up per column.
        checklist_by_label = {
            item.get("label"): item
            for item in (row.get("checklist") or [])
            if item.get("label")
        }

        for i, label in enumerate(checklist_labels):
            status_col = len(_CORE_COLUMNS) + 1 + (i * 3)
            item = checklist_by_label.get(label)
            if item is None:
                # This parcel's checklist doesn't include this item at all
                # (e.g. Osceola-only Option E row for a Pasco parcel).
                # Leave blank rather than fabricate a status.
                continue
            status = item.get("status")
            status_label = _STATUS_LABELS.get(status, status or "")
            cell = ws.cell(row=row_idx, column=status_col, value=status_label)
            fill = _STATUS_FILLS.get(status)
            if fill:
                cell.fill = fill
                cell.font = _STATUS_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # "Confirmed by" and "Date confirmed" left blank on purpose -- user fills in.

        # Notes column also left blank.

    # ---- freeze panes ----
    # Freeze row 1 AND the leftmost _FROZEN_COLUMN_COUNT identifier columns
    # so parcel_id/county/acres/owner stay visible while scrolling through
    # the pathway + checklist columns to the right. Cursor at "<letter>2"
    # tells Excel to freeze everything above + to the left of that cell.
    ws.freeze_panes = f"{get_column_letter(_FROZEN_COLUMN_COUNT + 1)}2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_filename(payload_rows: list[dict[str, Any]]) -> str:
    """
    Filename: `diligence_tracker_<counties>_<YYYY-MM-DD>.xlsx`, where
    counties is the sorted unique set of county_ids in the selection,
    joined by underscores. Capped at 4 counties (the pilot set) so it
    doesn't blow out to a novel-length filename if this ever handles
    dozens of counties -- extra get collapsed to a "+N" suffix.
    """
    counties = sorted({row.get("county_id") for row in payload_rows if row.get("county_id")})
    if len(counties) <= 4:
        county_slug = "_".join(counties) if counties else "no_county"
    else:
        county_slug = f"{'_'.join(counties[:4])}_plus{len(counties) - 4}"
    return f"enclave_candidates_{county_slug}_{date.today().isoformat()}.xlsx"
